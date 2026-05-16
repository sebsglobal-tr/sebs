#!/usr/bin/env python3
"""SEBS satışa hazırlık checklist → CSV + XLSX (Notion / Excel)."""
from __future__ import annotations

import csv
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "satis-hazirlik"

HEADERS = [
    "Öncelik",
    "ID",
    "Görev",
    "Sorumlu",
    "Tahmini gün",
    "Durum",
    "Not",
    "Kategori",
]

ROWS = [
    # P0
    ("P0", "L-01", "Gizlilik politikası sayfası", "Hukuk + Dev", "1", "Tamamlandı", "/gizlilik-politikasi.html — avukat incelemesi önerilir", "Yasal"),
    ("P0", "L-02", "Kullanım şartları (+ erken erişim maddesi)", "Hukuk + Dev", "1", "Tamamlandı", "/kullanim-sartlari.html", "Yasal"),
    ("P0", "L-03", "KVKK aydınlatma metni", "Hukuk + Dev", "1", "Tamamlandı", "/kvkk-aydinlatma.html", "Yasal"),
    ("P0", "L-04", "Footer / kayıt / iletişim yasal linkleri", "Dev", "0.5", "Tamamlandı", "/gizlilik, /kullanim-sartlari, /kvkk — avukat incelemesi ayrı", "Yasal"),
    ("P0", "C-01", "Checkout: şimdi / yakında + onay kutusu", "Dev", "1", "Tamamlandı", "pricing.html satın alma modalı", "Checkout"),
    ("P0", "C-02", "Production'da ücretsiz /api/purchase kapatma", "Dev", "0.5", "Tamamlandı", "NODE_ENV=production → 503 PAYMENT_PROVIDER_REQUIRED", "Checkout"),
    ("P0", "P-01", "Iyzico: sandbox hesap + API ürün tipi teyidi", "İş + Iyzico", "2-5", "Bekliyor", "Checkout Form API vs 3DS — yazılı teyit", "Ödeme"),
    ("P0", "P-02", "Backend: initialize + callback + sipariş tablosu", "Dev", "3-5", "Kısmen", "Kod hazır; IYZICO_API_KEY ile canlı test bekliyor", "Ödeme"),
    ("P0", "P-03", "Frontend: Ödemeye devam → Iyzico yönlendirme", "Dev", "1", "Kısmen", "pricing.html initialize çağırıyor", "Ödeme"),
    ("P0", "P-04", "Sunucu tarafı fiyat kataloğu", "Dev", "1", "Tamamlandı", "backend/lib/package-prices.js", "Ödeme"),
    ("P0", "P-05", "Entitlement yalnızca başarılı ödeme sonrası", "Dev", "1", "Kısmen", "Callback grant; prod /purchase kapalı", "Ödeme"),
    ("P0", "P-06", "Canlı smoke test: kayıt → ödeme → modül", "QA + Dev", "1", "Bekliyor", "P-02–05 sonrası", "QA"),
    # P1
    ("P1", "M-01", "Web Uygulama: paketten çıkar veya Yakında etiketi", "Ürün + Dev", "0.5", "Tamamlandı", "pricing yol haritası + erişim engeli", "Modül"),
    ("P1", "M-02", "temel-siber / ag-guvenligi yönlendirme", "Dev", "0.5", "Tamamlandı", "guncel-siber + network-guvenligi", "Modül"),
    ("P1", "M-03", "Simülasyon kataloğu rozet / link tutarlılığı", "Dev", "1", "Kısmen", "coming-soon kartları Yakında; bilgi notu var", "Simülasyon"),
    ("P1", "U-01", "Nav birleştirme (login/dashboard = marketing)", "Dev", "2", "Bekliyor", "", "UX"),
    ("P1", "U-02", "Eksik CSS/JS repoya al (styles.css, navigation.js)", "Dev", "1-2", "Kısmen", "styles/modules/enhancements.css stub; navigation.js mevcut", "Teknik"),
    ("P1", "A-01", "Modül erişim kontrolü (satın almayan → pricing)", "Dev", "2-3", "Kısmen", "modules.html + module-access-gate.js (doğrudan URL)", "Erişim"),
    ("P1", "I-01", "İade / destek süreci dokümantasyonu", "Operasyon", "0.5", "Tamamlandı", "kullanim-sartlari.html İade ve destek", "Operasyon"),
    # P2
    ("P2", "T-01", "Premium modüllerde UX standardizasyonu", "Dev", "5-10", "Bekliyor", "İşletim referans", "UX"),
    ("P2", "T-02", "Web modülü müfredatını yayınla", "İçerik + Dev", "5-15", "Bekliyor", "parts/ ~8k satır taslak", "İçerik"),
    ("P2", "T-03", "Bulut modülleri: premium motor veya satıştan çıkar", "Ürün", "2", "Bekliyor", "", "Modül"),
    ("P2", "T-04", "3 modül uçtan uca QA", "QA", "2", "Bekliyor", "Giriş + orta + ileri", "QA"),
    ("P2", "T-05", "Sertifika akışı doğrulama", "Dev + QA", "1", "Bekliyor", "", "QA"),
    ("P2", "T-06", "Sitemap / SEO temizliği", "Dev", "0.5", "Bekliyor", "temel-siber URL vb.", "SEO"),
]

SUMMARY_ROWS = [
    ("Özet", "", "P0 kalan (Iyzico + test)", "Dev", "8-12", "", "Erken erişim satış hedefi", ""),
    ("Özet", "", "P1 toplam", "Dev", "7-10", "", "", ""),
    ("Özet", "", "P2 toplam", "Dev + İçerik", "15-25", "", "İçerik dahil", ""),
    ("Özet", "", "Siber paket hazırlık (erken erişim)", "", "", "~%86-90", "Iyzico sandbox testi sonrası ~%90", ""),
]


def write_csv(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        w.writerows(ROWS)
        w.writerow([])
        w.writerow(["--- ÖZET ---"] + [""] * (len(HEADERS) - 1))
        w.writerows(SUMMARY_ROWS)


def write_xlsx(path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise SystemExit("openpyxl gerekli: python3 -m venv .venv && pip install openpyxl") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    status_colors = {
        "Tamamlandı": "D1FAE5",
        "Kısmen": "FEF3C7",
        "Bekliyor": "F1F5F9",
    }
    priority_colors = {"P0": "FEE2E2", "P1": "FFEDD5", "P2": "E0E7FF"}

    for r, row in enumerate(ROWS, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        pri = row[0]
        if pri in priority_colors:
            ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=priority_colors[pri])
        st = row[5]
        if st in status_colors:
            ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor=status_colors[st])

    start = len(ROWS) + 3
    ws.cell(row=start, column=1, value="ÖZET").font = Font(bold=True, size=12)
    for i, row in enumerate(SUMMARY_ROWS, start + 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=i, column=c, value=val)

    widths = [8, 8, 48, 14, 12, 14, 52, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS) + 1}"

    # Notion import sheet
    ws2 = wb.create_sheet("Notion rehber")
    guide = [
        ["Notion'a içe aktarma"],
        ["1. Notion → Import → CSV"],
        ["2. Dosya: satis-hazirlik-checklist-notion.csv"],
        ["3. Database olarak içe aktarın"],
        ["4. Öncelik / Durum / Sorumlu sütunlarını Select veya Status yapın"],
        ["5. Görev sütununu Title yapın"],
        [""],
        ["Alternatif: Bu Excel dosyasını Google Sheets'e yükleyip Notion Sync kullanın"],
    ]
    for r, line in enumerate(guide, 1):
        ws2.cell(row=r, column=1, value=line[0])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_notion_csv(path: Path) -> None:
    """Notion-friendly: Task name first column (Title), rest as properties."""
    notion_headers = [
        "Görev",
        "Öncelik",
        "ID",
        "Sorumlu",
        "Tahmini gün",
        "Durum",
        "Not",
        "Kategori",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(notion_headers)
        for row in ROWS:
            w.writerow([row[2], row[0], row[1], row[3], row[4], row[5], row[6], row[7]])


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    csv_path = OUT / "satis-hazirlik-checklist.csv"
    xlsx_path = OUT / "satis-hazirlik-checklist.xlsx"
    notion_csv = OUT / "satis-hazirlik-checklist-notion.csv"

    write_csv(csv_path)
    write_notion_csv(notion_csv)
    try:
        write_xlsx(xlsx_path)
        xlsx_ok = True
    except SystemExit:
        xlsx_ok = False
        print("XLSX atlandı (openpyxl yok). CSV dosyaları oluşturuldu.")

    print(f"CSV (Excel):  {csv_path}")
    print(f"CSV (Notion): {notion_csv}")
    if xlsx_ok:
        print(f"Excel:        {xlsx_path}")


if __name__ == "__main__":
    main()
